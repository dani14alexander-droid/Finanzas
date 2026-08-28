from pathlib import Path
from datetime import date, datetime, timedelta
from zoneinfo import ZoneInfo
from io import BytesIO
from math import ceil
import math
import calendar
import csv
import json
import os
import uuid

try:
    import psycopg
    from psycopg.rows import dict_row
except ImportError:
    psycopg = None
    dict_row = None

from flask import Flask, Response, redirect, render_template, request, send_file, url_for


app = Flask(__name__)

ZONA_HORARIA_CHILE = ZoneInfo("America/Santiago")

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
CSV_PATH = DATA_DIR / "finanzas.csv"
AUTOMATIZACIONES_PATH = DATA_DIR / "automatizaciones.csv"
DEUDAS_PATH = DATA_DIR / "deudas.csv"
PLANIFICACION_PATH = DATA_DIR / "planificacion.csv"
METAS_PATH = DATA_DIR / "metas.csv"
MACRO_CATEGORIAS_PATH = DATA_DIR / "macro_categorias.csv"
CATEGORIAS_PATH = DATA_DIR / "categorias.csv"
ENV_PATH = BASE_DIR / ".env"
LINK_PATH = BASE_DIR / "link.txt"
DB_LISTA = False
COLUMNAS = ["fecha", "tipo", "categoria", "descripcion", "monto", "ticket_deuda", "origen_deuda"]
AUTOMATIZACION_COLUMNAS = [
    "tipo",
    "categoria",
    "descripcion",
    "monto",
    "dia_mes",
    "activo",
    "ultimo_confirmado",
    "ticket_ultimo",
    "ultimo_anulado",
    "razon_anulado",
]
DEUDA_COLUMNAS = [
    "fecha",
    "tipo",
    "persona",
    "categoria",
    "descripcion",
    "monto",
    "estado",
    "fecha_pago",
    "modalidad",
    "cuotas_total",
    "cuotas_pagadas",
    "ticket_deuda",
]
PLANIFICACION_COLUMNAS = ["fecha", "tipo", "categoria", "descripcion", "monto"]
META_COLUMNAS = ["macro", "porcentaje"]
MACRO_CATEGORIA_COLUMNAS = ["tipo", "categoria", "macro"]
CATEGORIA_COLUMNAS = ["tipo", "categoria"]
CATEGORIA_SIN_ASIGNAR = "Sin categoría"
MACROS_ASIGNABLES = [
    "Ahorro principal",
    "Alimentacion",
    "Transporte",
    "Salidas y ocio",
    "Compras personales",
    "Fondo imprevistos",
]
METAS_PREDEFINIDAS = [
    ("Gastos fijos", 35),
    ("Ahorro principal", 15),
    ("Alimentacion", 15),
    ("Transporte", 10),
    ("Salidas y ocio", 10),
    ("Compras personales", 5),
    ("Fondo imprevistos", 5),
    ("Libre", 5),
]
TIPOS_VALIDOS = {"Ingreso", "Gasto", "Ahorro"}
TIPOS_AUTOMATIZACION = {"Gasto", "Ahorro"}
TIPOS_DEUDA = {"Me deben", "Debo"}
CATEGORIAS_PREDEFINIDAS = {
    "Ingreso": ["Sueldo", "Devolución de Impuestos", "Balance"],
    "Gasto": [
        "Deuda",
        "Comida",
        "Arriendo",
        "Salud e imagen",
        "Familia",
        "Entretenimiento",
        "Ropa",
        "Mascotas",
        "Regalo",
        "Giro",
        "Prestamo",
        "Educación",
        "Balance",
        "Transporte",
    ],
    "Ahorro": ["Ahorro", "Inversión", "Fondo de emergencia", "Balance"],
    "Me deben": ["Prestamo", "Adelanto", "Compra compartida"],
    "Debo": ["Prestamo", "Tarjeta", "Adelanto"],
}
MESES = [
    "Enero",
    "Febrero",
    "Marzo",
    "Abril",
    "Mayo",
    "Junio",
    "Julio",
    "Agosto",
    "Septiembre",
    "Octubre",
    "Noviembre",
    "Diciembre",
]


def cargar_entorno_local():
    if ENV_PATH.exists():
        for linea in ENV_PATH.read_text(encoding="utf-8").splitlines():
            linea = linea.strip()
            if not linea or linea.startswith("#") or "=" not in linea:
                continue
            clave, valor = linea.split("=", 1)
            os.environ.setdefault(clave.strip(), valor.strip().strip('"').strip("'"))
    if not os.getenv("DATABASE_URL") and LINK_PATH.exists():
        os.environ["DATABASE_URL"] = LINK_PATH.read_text(encoding="utf-8").strip()


cargar_entorno_local()


def database_url():
    url = os.getenv("DATABASE_URL", "").strip()
    if url and "sslmode=" not in url:
        separador = "&" if "?" in url else "?"
        url = f"{url}{separador}sslmode=require"
    return url


def usar_base_datos():
    return bool(database_url() and psycopg)


def conectar_db():
    return psycopg.connect(database_url(), row_factory=dict_row)


def error_base_datos(error):
    app.logger.error("La base de datos no esta disponible: %s", error)
    template = app.jinja_env.get_template("base_datos_no_disponible.html")
    return template.render(), 503


if psycopg:
    app.register_error_handler(psycopg.Error, error_base_datos)


def asegurar_db():
    global DB_LISTA
    if not usar_base_datos():
        return
    if DB_LISTA:
        return
    with conectar_db() as conexion:
        with conexion.cursor() as cursor:
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS movimientos (
                    id BIGSERIAL PRIMARY KEY,
                    fecha TEXT NOT NULL DEFAULT '',
                    tipo TEXT NOT NULL DEFAULT '',
                    categoria TEXT NOT NULL DEFAULT '',
                    descripcion TEXT NOT NULL DEFAULT '',
                    monto DOUBLE PRECISION NOT NULL DEFAULT 0,
                    ticket_deuda TEXT NOT NULL DEFAULT '',
                    origen_deuda TEXT NOT NULL DEFAULT ''
                )
                """
            )
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS macro_categorias (
                    tipo TEXT NOT NULL DEFAULT '',
                    categoria TEXT NOT NULL DEFAULT '',
                    macro TEXT NOT NULL DEFAULT '',
                    PRIMARY KEY (tipo, categoria)
                )
                """
            )
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS categorias (
                    tipo TEXT NOT NULL DEFAULT '',
                    categoria TEXT NOT NULL DEFAULT '',
                    PRIMARY KEY (tipo, categoria)
                )
                """
            )
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS automatizaciones (
                    id BIGSERIAL PRIMARY KEY,
                    tipo TEXT NOT NULL DEFAULT '',
                    categoria TEXT NOT NULL DEFAULT '',
                    descripcion TEXT NOT NULL DEFAULT '',
                    monto DOUBLE PRECISION NOT NULL DEFAULT 0,
                    dia_mes INTEGER NOT NULL DEFAULT 1,
                    activo BOOLEAN NOT NULL DEFAULT TRUE,
                    ultimo_confirmado TEXT NOT NULL DEFAULT '',
                    ticket_ultimo TEXT NOT NULL DEFAULT '',
                    ultimo_anulado TEXT NOT NULL DEFAULT '',
                    razon_anulado TEXT NOT NULL DEFAULT ''
                )
                """
            )
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS deudas (
                    id BIGSERIAL PRIMARY KEY,
                    fecha TEXT NOT NULL DEFAULT '',
                    tipo TEXT NOT NULL DEFAULT '',
                    persona TEXT NOT NULL DEFAULT '',
                    categoria TEXT NOT NULL DEFAULT '',
                    descripcion TEXT NOT NULL DEFAULT '',
                    monto DOUBLE PRECISION NOT NULL DEFAULT 0,
                    estado TEXT NOT NULL DEFAULT '',
                    fecha_pago TEXT NOT NULL DEFAULT '',
                    modalidad TEXT NOT NULL DEFAULT '',
                    cuotas_total INTEGER NOT NULL DEFAULT 1,
                    cuotas_pagadas INTEGER NOT NULL DEFAULT 0,
                    ticket_deuda TEXT NOT NULL DEFAULT ''
                )
                """
            )
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS planificacion (
                    id BIGSERIAL PRIMARY KEY,
                    fecha TEXT NOT NULL DEFAULT '',
                    tipo TEXT NOT NULL DEFAULT '',
                    categoria TEXT NOT NULL DEFAULT '',
                    descripcion TEXT NOT NULL DEFAULT '',
                    monto DOUBLE PRECISION NOT NULL DEFAULT 0
                )
                """
            )
            cursor.execute("ALTER TABLE movimientos ADD COLUMN IF NOT EXISTS ticket_deuda TEXT NOT NULL DEFAULT ''")
            cursor.execute("ALTER TABLE movimientos ADD COLUMN IF NOT EXISTS origen_deuda TEXT NOT NULL DEFAULT ''")
            cursor.execute("ALTER TABLE deudas ADD COLUMN IF NOT EXISTS modalidad TEXT NOT NULL DEFAULT ''")
            cursor.execute("ALTER TABLE deudas ADD COLUMN IF NOT EXISTS cuotas_total INTEGER NOT NULL DEFAULT 1")
            cursor.execute("ALTER TABLE deudas ADD COLUMN IF NOT EXISTS cuotas_pagadas INTEGER NOT NULL DEFAULT 0")
            cursor.execute("ALTER TABLE deudas ADD COLUMN IF NOT EXISTS ticket_deuda TEXT NOT NULL DEFAULT ''")
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS metas (
                    macro TEXT PRIMARY KEY,
                    porcentaje DOUBLE PRECISION NOT NULL DEFAULT 0
                )
                """
            )
            cursor.execute("SELECT COUNT(*) AS total FROM categorias")
            if cursor.fetchone()["total"] == 0:
                cursor.executemany(
                    "INSERT INTO categorias (tipo, categoria) VALUES (%s, %s) ON CONFLICT DO NOTHING",
                    [
                        (tipo, categoria)
                        for tipo in TIPOS_VALIDOS
                        for categoria in CATEGORIAS_PREDEFINIDAS.get(tipo, [])
                    ],
                )
            cursor.executemany(
                "INSERT INTO categorias (tipo, categoria) VALUES (%s, %s) ON CONFLICT DO NOTHING",
                [(tipo, CATEGORIA_SIN_ASIGNAR) for tipo in TIPOS_VALIDOS],
            )
            cursor.execute(
                """
                INSERT INTO categorias (tipo, categoria)
                SELECT DISTINCT tipo, categoria FROM movimientos
                WHERE tipo IN ('Ingreso', 'Gasto', 'Ahorro') AND BTRIM(categoria) <> ''
                ON CONFLICT DO NOTHING
                """
            )
    DB_LISTA = True


def asegurar_columnas_csv(path, columnas):
    if not path.exists():
        return
    with path.open(newline="", encoding="utf-8") as archivo:
        reader = csv.DictReader(archivo)
        if reader.fieldnames == columnas:
            return
        filas = list(reader)
    with path.open("w", newline="", encoding="utf-8") as archivo:
        writer = csv.DictWriter(archivo, fieldnames=columnas)
        writer.writeheader()
        for fila in filas:
            writer.writerow({columna: fila.get(columna, "") for columna in columnas})


def asegurar_csv():
    DATA_DIR.mkdir(exist_ok=True)
    if not CSV_PATH.exists():
        with CSV_PATH.open("w", newline="", encoding="utf-8") as archivo:
            writer = csv.DictWriter(archivo, fieldnames=COLUMNAS)
            writer.writeheader()
    if not AUTOMATIZACIONES_PATH.exists():
        with AUTOMATIZACIONES_PATH.open("w", newline="", encoding="utf-8") as archivo:
            writer = csv.DictWriter(archivo, fieldnames=AUTOMATIZACION_COLUMNAS)
            writer.writeheader()
    if not DEUDAS_PATH.exists():
        with DEUDAS_PATH.open("w", newline="", encoding="utf-8") as archivo:
            writer = csv.DictWriter(archivo, fieldnames=DEUDA_COLUMNAS)
            writer.writeheader()
    if not PLANIFICACION_PATH.exists():
        with PLANIFICACION_PATH.open("w", newline="", encoding="utf-8") as archivo:
            writer = csv.DictWriter(archivo, fieldnames=PLANIFICACION_COLUMNAS)
            writer.writeheader()
    if not METAS_PATH.exists():
        with METAS_PATH.open("w", newline="", encoding="utf-8") as archivo:
            writer = csv.DictWriter(archivo, fieldnames=META_COLUMNAS)
            writer.writeheader()
            for macro, porcentaje in METAS_PREDEFINIDAS:
                writer.writerow({"macro": macro, "porcentaje": porcentaje})
    if not MACRO_CATEGORIAS_PATH.exists():
        with MACRO_CATEGORIAS_PATH.open("w", newline="", encoding="utf-8") as archivo:
            writer = csv.DictWriter(archivo, fieldnames=MACRO_CATEGORIA_COLUMNAS)
            writer.writeheader()
    if not CATEGORIAS_PATH.exists():
        categorias_iniciales = []
        for tipo in TIPOS_VALIDOS:
            for categoria in [*CATEGORIAS_PREDEFINIDAS.get(tipo, []), CATEGORIA_SIN_ASIGNAR]:
                categorias_iniciales.append({"tipo": tipo, "categoria": categoria})
        if CSV_PATH.exists():
            with CSV_PATH.open(newline="", encoding="utf-8") as archivo:
                for fila in csv.DictReader(archivo):
                    if fila.get("tipo") in TIPOS_VALIDOS and fila.get("categoria", "").strip():
                        categorias_iniciales.append(
                            {"tipo": fila["tipo"], "categoria": fila["categoria"].strip()}
                        )
        unicas = {
            (item["tipo"], item["categoria"].lower()): item
            for item in categorias_iniciales
        }
        with CATEGORIAS_PATH.open("w", newline="", encoding="utf-8") as archivo:
            writer = csv.DictWriter(archivo, fieldnames=CATEGORIA_COLUMNAS)
            writer.writeheader()
            writer.writerows(unicas.values())
    asegurar_columnas_csv(CSV_PATH, COLUMNAS)
    asegurar_columnas_csv(DEUDAS_PATH, DEUDA_COLUMNAS)


def leer_movimientos():
    if usar_base_datos():
        asegurar_db()
        with conectar_db() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute(
                    "SELECT fecha, tipo, categoria, descripcion, monto, ticket_deuda, origen_deuda FROM movimientos ORDER BY id"
                )
                movimientos = []
                for indice, fila in enumerate(cursor.fetchall()):
                    movimiento = {columna: fila.get(columna) or "" for columna in COLUMNAS}
                    movimiento["monto"] = float(movimiento["monto"] or 0)
                    movimiento["id"] = indice
                    movimientos.append(movimiento)
                return movimientos

    asegurar_csv()
    with CSV_PATH.open(newline="", encoding="utf-8") as archivo:
        reader = csv.DictReader(archivo)
        movimientos = []
        for indice, fila in enumerate(reader):
            movimiento = {columna: fila.get(columna, "") for columna in COLUMNAS}
            try:
                movimiento["monto"] = float(movimiento["monto"] or 0)
            except ValueError:
                movimiento["monto"] = 0
            movimiento["id"] = indice
            movimientos.append(movimiento)
    return movimientos


def escribir_movimientos(movimientos):
    if usar_base_datos():
        asegurar_db()
        with conectar_db() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("TRUNCATE movimientos RESTART IDENTITY")
                cursor.executemany(
                    """
                    INSERT INTO movimientos (fecha, tipo, categoria, descripcion, monto, ticket_deuda, origen_deuda)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """,
                    [
                        (
                            movimiento.get("fecha", ""),
                            movimiento.get("tipo", ""),
                            movimiento.get("categoria", ""),
                            movimiento.get("descripcion", ""),
                            float(movimiento.get("monto") or 0),
                            movimiento.get("ticket_deuda", ""),
                            movimiento.get("origen_deuda", ""),
                        )
                        for movimiento in movimientos
                    ],
                )
        return

    asegurar_csv()
    with CSV_PATH.open("w", newline="", encoding="utf-8") as archivo:
        writer = csv.DictWriter(archivo, fieldnames=COLUMNAS)
        writer.writeheader()
        for movimiento in movimientos:
            writer.writerow({columna: movimiento.get(columna, "") for columna in COLUMNAS})


def guardar_movimiento(movimiento):
    if usar_base_datos():
        asegurar_db()
        with conectar_db() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute(
                    """
                    INSERT INTO movimientos (fecha, tipo, categoria, descripcion, monto, ticket_deuda, origen_deuda)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        movimiento.get("fecha", ""),
                        movimiento.get("tipo", ""),
                        movimiento.get("categoria", ""),
                        movimiento.get("descripcion", ""),
                        float(movimiento.get("monto") or 0),
                        movimiento.get("ticket_deuda", ""),
                        movimiento.get("origen_deuda", ""),
                    ),
                )
        return

    asegurar_csv()
    with CSV_PATH.open("a", newline="", encoding="utf-8") as archivo:
        writer = csv.DictWriter(archivo, fieldnames=COLUMNAS)
        writer.writerow(movimiento)


def leer_automatizaciones():
    if usar_base_datos():
        asegurar_db()
        with conectar_db() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT tipo, categoria, descripcion, monto, dia_mes, activo,
                           ultimo_confirmado, ticket_ultimo, ultimo_anulado, razon_anulado
                    FROM automatizaciones
                    ORDER BY id
                    """
                )
                automatizaciones = []
                for indice, fila in enumerate(cursor.fetchall()):
                    item = {columna: fila.get(columna) or "" for columna in AUTOMATIZACION_COLUMNAS}
                    item["monto"] = float(item["monto"] or 0)
                    item["dia_mes"] = int(item["dia_mes"] or 1)
                    item["activo"] = bool(item["activo"])
                    item["id"] = indice
                    automatizaciones.append(item)
                return automatizaciones

    asegurar_csv()
    with AUTOMATIZACIONES_PATH.open(newline="", encoding="utf-8") as archivo:
        reader = csv.DictReader(archivo)
        automatizaciones = []
        for indice, fila in enumerate(reader):
            item = {columna: fila.get(columna, "") for columna in AUTOMATIZACION_COLUMNAS}
            try:
                item["monto"] = float(item["monto"] or 0)
            except ValueError:
                item["monto"] = 0
            try:
                item["dia_mes"] = int(item["dia_mes"] or 1)
            except ValueError:
                item["dia_mes"] = 1
            item["id"] = indice
            item["activo"] = item["activo"] != "No"
            automatizaciones.append(item)
    return automatizaciones


def escribir_automatizaciones(automatizaciones):
    if usar_base_datos():
        asegurar_db()
        with conectar_db() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("TRUNCATE automatizaciones RESTART IDENTITY")
                cursor.executemany(
                    """
                    INSERT INTO automatizaciones (
                        tipo, categoria, descripcion, monto, dia_mes, activo,
                        ultimo_confirmado, ticket_ultimo, ultimo_anulado, razon_anulado
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    [
                        (
                            item.get("tipo", ""),
                            item.get("categoria", ""),
                            item.get("descripcion", ""),
                            float(item.get("monto") or 0),
                            int(item.get("dia_mes") or 1),
                            bool(item.get("activo", True)),
                            item.get("ultimo_confirmado", ""),
                            item.get("ticket_ultimo", ""),
                            item.get("ultimo_anulado", ""),
                            item.get("razon_anulado", ""),
                        )
                        for item in automatizaciones
                    ],
                )
        return

    asegurar_csv()
    with AUTOMATIZACIONES_PATH.open("w", newline="", encoding="utf-8") as archivo:
        writer = csv.DictWriter(archivo, fieldnames=AUTOMATIZACION_COLUMNAS)
        writer.writeheader()
        for item in automatizaciones:
            fila = {columna: item.get(columna, "") for columna in AUTOMATIZACION_COLUMNAS}
            fila["activo"] = "Si" if item.get("activo", True) else "No"
            writer.writerow(fila)


def leer_deudas():
    if usar_base_datos():
        asegurar_db()
        with conectar_db() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT fecha, tipo, persona, categoria, descripcion, monto, estado, fecha_pago,
                           modalidad, cuotas_total, cuotas_pagadas, ticket_deuda
                    FROM deudas
                    ORDER BY id
                    """
                )
                deudas = []
                for indice, fila in enumerate(cursor.fetchall()):
                    item = {columna: fila.get(columna) or "" for columna in DEUDA_COLUMNAS}
                    item["monto"] = float(item["monto"] or 0)
                    item["cuotas_total"] = int(item["cuotas_total"] or 1)
                    item["cuotas_pagadas"] = int(item["cuotas_pagadas"] or 0)
                    item["id"] = indice
                    deudas.append(item)
                return deudas

    asegurar_csv()
    with DEUDAS_PATH.open(newline="", encoding="utf-8") as archivo:
        reader = csv.DictReader(archivo)
        deudas = []
        for indice, fila in enumerate(reader):
            item = {columna: fila.get(columna, "") for columna in DEUDA_COLUMNAS}
            try:
                item["monto"] = float(item["monto"] or 0)
            except ValueError:
                item["monto"] = 0
            item["id"] = indice
            try:
                item["cuotas_total"] = max(int(item.get("cuotas_total") or 1), 1)
            except ValueError:
                item["cuotas_total"] = 1
            try:
                item["cuotas_pagadas"] = max(int(item.get("cuotas_pagadas") or 0), 0)
            except ValueError:
                item["cuotas_pagadas"] = 0
            deudas.append(item)
    return deudas


def leer_categorias():
    if usar_base_datos():
        asegurar_db()
        with conectar_db() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT tipo, categoria FROM categorias ORDER BY tipo, LOWER(categoria)")
                return [dict(fila) for fila in cursor.fetchall()]
    asegurar_csv()
    with CATEGORIAS_PATH.open(newline="", encoding="utf-8") as archivo:
        return [dict(fila) for fila in csv.DictReader(archivo)]


def escribir_categorias(categorias):
    unicas = {}
    for item in categorias:
        tipo = item.get("tipo", "")
        categoria = item.get("categoria", "").strip()
        if tipo in TIPOS_VALIDOS and categoria:
            unicas[(tipo, categoria.lower())] = {"tipo": tipo, "categoria": categoria}
    categorias_limpias = sorted(
        unicas.values(), key=lambda item: (item["tipo"], item["categoria"].lower())
    )
    if usar_base_datos():
        asegurar_db()
        with conectar_db() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("TRUNCATE categorias")
                cursor.executemany(
                    "INSERT INTO categorias (tipo, categoria) VALUES (%s, %s)",
                    [(item["tipo"], item["categoria"]) for item in categorias_limpias],
                )
        return
    asegurar_csv()
    with CATEGORIAS_PATH.open("w", newline="", encoding="utf-8") as archivo:
        writer = csv.DictWriter(archivo, fieldnames=CATEGORIA_COLUMNAS)
        writer.writeheader()
        writer.writerows(categorias_limpias)


def categorias_existentes():
    categorias = {"Todas": {}}
    tipos = TIPOS_VALIDOS | TIPOS_AUTOMATIZACION | TIPOS_DEUDA
    for tipo in tipos:
        categorias[tipo] = {}

    def agregar_categoria(tipo, categoria):
        categoria = categoria.strip()
        if not categoria:
            return
        categorias.setdefault(tipo, {})
        categorias[tipo].setdefault(categoria.lower(), categoria)
        categorias["Todas"].setdefault(categoria.lower(), categoria)

    for tipo, valores in CATEGORIAS_PREDEFINIDAS.items():
        if tipo in TIPOS_VALIDOS:
            continue
        for categoria in valores:
            agregar_categoria(tipo, categoria)

    for item in leer_categorias():
        agregar_categoria(item.get("tipo", ""), item.get("categoria", ""))
    for coleccion in (leer_movimientos(), leer_automatizaciones(), leer_planificaciones()):
        for item in coleccion:
            agregar_categoria(item.get("tipo", ""), item.get("categoria", ""))
    for item in leer_deudas():
        agregar_categoria(item.get("tipo", ""), item.get("categoria", ""))

    return {
        tipo: sorted(valores.values(), key=str.lower)
        for tipo, valores in categorias.items()
    }


def lista_categoria_para_tipo(tipo):
    ids = {
        "Ingreso": "categorias-ingreso",
        "Gasto": "categorias-gasto",
        "Ahorro": "categorias-ahorro",
        "Me deben": "categorias-me-deben",
        "Debo": "categorias-debo",
    }
    return ids.get(tipo, "categorias-todas")


@app.context_processor
def inyectar_categorias():
    return {
        "categorias_existentes": categorias_existentes(),
        "lista_categoria_para_tipo": lista_categoria_para_tipo,
    }
def escribir_deudas(deudas):
    if usar_base_datos():
        asegurar_db()
        with conectar_db() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("TRUNCATE deudas RESTART IDENTITY")
                cursor.executemany(
                    """
                    INSERT INTO deudas (
                        fecha, tipo, persona, categoria, descripcion, monto, estado, fecha_pago,
                        modalidad, cuotas_total, cuotas_pagadas, ticket_deuda
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    [
                        (
                            item.get("fecha", ""),
                            item.get("tipo", ""),
                            item.get("persona", ""),
                            item.get("categoria", ""),
                            item.get("descripcion", ""),
                            float(item.get("monto") or 0),
                            item.get("estado", ""),
                            item.get("fecha_pago", ""),
                            item.get("modalidad", ""),
                            int(item.get("cuotas_total") or 1),
                            int(item.get("cuotas_pagadas") or 0),
                            item.get("ticket_deuda", ""),
                        )
                        for item in deudas
                    ],
                )
        return

    asegurar_csv()
    with DEUDAS_PATH.open("w", newline="", encoding="utf-8") as archivo:
        writer = csv.DictWriter(archivo, fieldnames=DEUDA_COLUMNAS)
        writer.writeheader()
        for item in deudas:
            writer.writerow({columna: item.get(columna, "") for columna in DEUDA_COLUMNAS})


def leer_planificaciones():
    if usar_base_datos():
        asegurar_db()
        with conectar_db() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute(
                    "SELECT fecha, tipo, categoria, descripcion, monto FROM planificacion ORDER BY id"
                )
                planificaciones = []
                for indice, fila in enumerate(cursor.fetchall()):
                    item = {columna: fila.get(columna) or "" for columna in PLANIFICACION_COLUMNAS}
                    item["monto"] = float(item["monto"] or 0)
                    item["id"] = indice
                    planificaciones.append(item)
                return planificaciones

    asegurar_csv()
    with PLANIFICACION_PATH.open(newline="", encoding="utf-8") as archivo:
        reader = csv.DictReader(archivo)
        planificaciones = []
        for indice, fila in enumerate(reader):
            item = {columna: fila.get(columna, "") for columna in PLANIFICACION_COLUMNAS}
            try:
                item["monto"] = float(item["monto"] or 0)
            except ValueError:
                item["monto"] = 0
            item["id"] = indice
            planificaciones.append(item)
    return planificaciones


def escribir_planificaciones(planificaciones):
    if usar_base_datos():
        asegurar_db()
        with conectar_db() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("TRUNCATE planificacion RESTART IDENTITY")
                cursor.executemany(
                    """
                    INSERT INTO planificacion (fecha, tipo, categoria, descripcion, monto)
                    VALUES (%s, %s, %s, %s, %s)
                    """,
                    [
                        (
                            item.get("fecha", ""),
                            item.get("tipo", ""),
                            item.get("categoria", ""),
                            item.get("descripcion", ""),
                            float(item.get("monto") or 0),
                        )
                        for item in planificaciones
                    ],
                )
        return

    asegurar_csv()
    with PLANIFICACION_PATH.open("w", newline="", encoding="utf-8") as archivo:
        writer = csv.DictWriter(archivo, fieldnames=PLANIFICACION_COLUMNAS)
        writer.writeheader()
        for item in planificaciones:
            writer.writerow({columna: item.get(columna, "") for columna in PLANIFICACION_COLUMNAS})


def guardar_planificacion(item):
    if usar_base_datos():
        asegurar_db()
        with conectar_db() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute(
                    """
                    INSERT INTO planificacion (fecha, tipo, categoria, descripcion, monto)
                    VALUES (%s, %s, %s, %s, %s)
                    """,
                    (
                        item.get("fecha", ""),
                        item.get("tipo", ""),
                        item.get("categoria", ""),
                        item.get("descripcion", ""),
                        float(item.get("monto") or 0),
                    ),
                )
        return

    asegurar_csv()
    with PLANIFICACION_PATH.open("a", newline="", encoding="utf-8") as archivo:
        writer = csv.DictWriter(archivo, fieldnames=PLANIFICACION_COLUMNAS)
        writer.writerow(item)


def leer_metas():
    if usar_base_datos():
        asegurar_db()
        with conectar_db() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT macro, porcentaje FROM metas")
                guardadas = {fila["macro"]: float(fila["porcentaje"] or 0) for fila in cursor.fetchall()}
    else:
        asegurar_csv()
        with METAS_PATH.open(newline="", encoding="utf-8") as archivo:
            guardadas = {
                fila.get("macro", ""): float(fila.get("porcentaje") or 0)
                for fila in csv.DictReader(archivo)
            }
    return [
        {"macro": macro, "porcentaje": guardadas.get(macro, porcentaje)}
        for macro, porcentaje in METAS_PREDEFINIDAS
    ]


def escribir_metas(metas):
    if usar_base_datos():
        asegurar_db()
        with conectar_db() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("TRUNCATE metas")
                cursor.executemany(
                    "INSERT INTO metas (macro, porcentaje) VALUES (%s, %s)",
                    [(item["macro"], float(item["porcentaje"])) for item in metas],
                )
        return
    asegurar_csv()
    with METAS_PATH.open("w", newline="", encoding="utf-8") as archivo:
        writer = csv.DictWriter(archivo, fieldnames=META_COLUMNAS)
        writer.writeheader()
        writer.writerows(metas)


def leer_macro_categorias():
    if usar_base_datos():
        asegurar_db()
        with conectar_db() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("SELECT tipo, categoria, macro FROM macro_categorias")
                return [dict(fila) for fila in cursor.fetchall()]
    asegurar_csv()
    with MACRO_CATEGORIAS_PATH.open(newline="", encoding="utf-8") as archivo:
        return [dict(fila) for fila in csv.DictReader(archivo)]


def escribir_macro_categorias(asignaciones):
    if usar_base_datos():
        asegurar_db()
        with conectar_db() as conexion:
            with conexion.cursor() as cursor:
                cursor.execute("TRUNCATE macro_categorias")
                cursor.executemany(
                    "INSERT INTO macro_categorias (tipo, categoria, macro) VALUES (%s, %s, %s)",
                    [(item["tipo"], item["categoria"], item["macro"]) for item in asignaciones],
                )
        return
    asegurar_csv()
    with MACRO_CATEGORIAS_PATH.open("w", newline="", encoding="utf-8") as archivo:
        writer = csv.DictWriter(archivo, fieldnames=MACRO_CATEGORIA_COLUMNAS)
        writer.writeheader()
        writer.writerows(asignaciones)


def asignaciones_macro_dict():
    return {
        (item["tipo"], item["categoria"].strip().lower()): item["macro"]
        for item in leer_macro_categorias()
        if item.get("macro") in MACROS_ASIGNABLES
    }


def fecha_hoy_chile():
    return datetime.now(ZONA_HORARIA_CHILE).date()


def periodo_actual():
    return clave_ciclo(fecha_hoy_chile())


def proxima_fecha_mensual(dia_mes):
    hoy = fecha_hoy_chile()
    ultimo_dia = calendar.monthrange(hoy.year, hoy.month)[1]
    dia = min(max(int(dia_mes), 1), ultimo_dia)
    return date(hoy.year, hoy.month, dia).isoformat()


def penultimo_dia_habil_mes(hoy=None):
    hoy = hoy or fecha_hoy_chile()
    ultimo_dia = calendar.monthrange(hoy.year, hoy.month)[1]
    dias_habiles = [
        date(hoy.year, hoy.month, dia)
        for dia in range(1, ultimo_dia + 1)
        if date(hoy.year, hoy.month, dia).weekday() < 5
    ]
    if len(dias_habiles) >= 2:
        return dias_habiles[-2].isoformat()
    return dias_habiles[-1].isoformat()


def fecha_sueldo_esperada(hoy=None):
    return fecha_movimiento(penultimo_dia_habil_mes(hoy))


def sumar_meses(fecha, meses):
    mes = fecha.month - 1 + meses
    anio = fecha.year + mes // 12
    mes = mes % 12 + 1
    dia = min(fecha.day, calendar.monthrange(anio, mes)[1])
    return date(anio, mes, dia)


def es_sueldo(movimiento):
    if movimiento["tipo"] != "Ingreso":
        return False
    categoria = movimiento["categoria"].strip().lower()
    descripcion = movimiento["descripcion"].strip().lower()
    return categoria == "sueldo" or descripcion == "sueldo"


def ultimo_sueldo(movimientos):
    sueldos = [
        (fecha, item)
        for item in movimientos
        if es_sueldo(item) and (fecha := fecha_movimiento(item["fecha"]))
    ]
    if not sueldos:
        return None
    return max(sueldos, key=lambda item: item[0])


def ultimo_sueldo_antes_de(movimientos, fecha_limite):
    sueldos = [
        (fecha, item)
        for item in movimientos
        if es_sueldo(item)
        and (fecha := fecha_movimiento(item["fecha"]))
        and fecha < fecha_limite
    ]
    if not sueldos:
        return None
    return max(sueldos, key=lambda item: item[0])


def clave_ciclo(fecha, movimientos=None):
    movimientos = movimientos if movimientos is not None else leer_movimientos()
    sueldos = [
        fecha_sueldo
        for item in movimientos
        if es_sueldo(item)
        and (fecha_sueldo := fecha_movimiento(item.get("fecha", "")))
        and fecha_sueldo <= fecha
    ]
    if sueldos:
        return max(sueldos).isoformat()
    return f"{fecha.year}-{fecha.month:02d}"


def periodo_movimiento_ciclo(valor, movimientos=None):
    fecha = fecha_movimiento(valor)
    if not fecha:
        return ""
    return clave_ciclo(fecha, movimientos)


def pago_deuda_genera_movimiento(deuda, fecha_pago, movimientos):
    tiene_movimiento_origen = deuda.get("tipo") == "Me deben" or (
        deuda.get("tipo") == "Debo" and deuda.get("modalidad") == "Prestamo"
    )
    if not tiene_movimiento_origen:
        return True

    fecha_deuda = fecha_movimiento(deuda.get("fecha", ""))
    fecha_pagada = fecha_movimiento(fecha_pago)
    if not fecha_deuda or not fecha_pagada:
        return True

    ciclo_deuda = clave_ciclo(fecha_deuda, movimientos)
    ciclo_pago = clave_ciclo(fecha_pagada, movimientos)
    return ciclo_deuda != ciclo_pago


def monto_cuota(deuda):
    cuotas = max(int(deuda.get("cuotas_total") or 1), 1)
    return float(deuda.get("monto") or 0) / cuotas


def eliminar_movimientos_deuda(ticket, origen=None):
    if not ticket:
        return
    movimientos = leer_movimientos()
    filtrados = [
        item
        for item in movimientos
        if not (
            item.get("ticket_deuda") == ticket
            and (origen is None or item.get("origen_deuda") == origen)
        )
    ]
    if len(filtrados) != len(movimientos):
        escribir_movimientos(filtrados)


def saldo_ciclo_anterior(movimientos, fecha_inicio):
    mes_anterior = sumar_meses(date(fecha_inicio.year, fecha_inicio.month, 1), -1)
    inicio_anterior = fecha_movimiento(penultimo_dia_habil_mes(mes_anterior))
    fin_anterior = fecha_inicio - timedelta(days=1)
    saldo = 0
    for item in movimientos:
        fecha = fecha_movimiento(item.get("fecha", ""))
        if not fecha or not (inicio_anterior <= fecha <= fin_anterior):
            continue
        monto = float(item.get("monto") or 0)
        if item.get("tipo") == "Ingreso":
            saldo += monto
        elif item.get("tipo") in {"Gasto", "Ahorro"}:
            saldo -= monto
    return saldo


def ciclo_anterior_tiene_informacion(movimientos, inicio):
    mes_anterior = sumar_meses(date(inicio.year, inicio.month, 1), -1)
    inicio_anterior = fecha_movimiento(penultimo_dia_habil_mes(mes_anterior))
    fin_anterior = inicio - timedelta(days=1)
    return any(
        (fecha := fecha_movimiento(item.get("fecha", "")))
        and inicio_anterior <= fecha <= fin_anterior
        for item in movimientos
    )


def etiqueta_ciclo(clave):
    fecha = fecha_movimiento(clave)
    if fecha:
        proximo_mes = sumar_meses(fecha, 1)
        fecha_fin = fecha_movimiento(penultimo_dia_habil_mes(proximo_mes)) - timedelta(days=1)
        mes, anio = mes_dominante(fecha, fecha_fin)
        return f"Ciclo {MESES[mes - 1]} {anio}"
    try:
        anio, mes = clave.split("-")[:2]
        return f"Ciclo {MESES[int(mes) - 1]} {anio}"
    except (ValueError, IndexError):
        return "Ciclo actual"


def fecha_corta(valor):
    if hasattr(valor, "strftime"):
        return valor.strftime("%d-%m-%Y")
    fecha = fecha_movimiento(str(valor))
    if fecha:
        return fecha.strftime("%d-%m-%Y")
    return ""


def mes_dominante(inicio, fin):
    dias_por_mes = {}
    actual = inicio
    while actual <= fin:
        clave = (actual.year, actual.month)
        dias_por_mes[clave] = dias_por_mes.get(clave, 0) + 1
        actual += timedelta(days=1)
    anio, mes = max(dias_por_mes, key=lambda clave: dias_por_mes[clave])
    return mes, anio


def rango_ciclo(clave, movimientos):
    fecha_inicio = fecha_movimiento(clave)
    if fecha_inicio:
        proximo_mes = sumar_meses(fecha_inicio, 1)
        fecha_fin = fecha_movimiento(penultimo_dia_habil_mes(proximo_mes))
        return fecha_inicio, fecha_fin - timedelta(days=1)

    try:
        anio, mes = [int(parte) for parte in clave.split("-")[:2]]
    except (ValueError, IndexError):
        return rango_dashboard(fecha_hoy_chile(), movimientos)

    inicio = date(anio, mes, 1)
    fin = date(anio, mes, calendar.monthrange(anio, mes)[1])
    return inicio, fin


def opciones_ciclos(movimientos):
    claves = {
        fecha.isoformat()
        for item in movimientos
        if es_sueldo(item) and (fecha := fecha_movimiento(item.get("fecha", "")))
    }
    claves.add(clave_ciclo(fecha_hoy_chile(), movimientos))

    # Include missing salary periods so months without activity remain visible.
    fechas_ciclo = sorted(fecha_movimiento(clave) for clave in claves if fecha_movimiento(clave))
    if fechas_ciclo:
        cursor = date(fechas_ciclo[0].year, fechas_ciclo[0].month, 1)
        ultimo = date(fechas_ciclo[-1].year, fechas_ciclo[-1].month, 1)
        while cursor <= ultimo:
            claves.add(penultimo_dia_habil_mes(cursor))
            cursor = sumar_meses(cursor, 1)

    if not claves:
        claves = {
            f"{fecha.year}-{fecha.month:02d}"
            for item in movimientos
            if (fecha := fecha_movimiento(item.get("fecha", "")))
        }

    opciones = []
    for clave in sorted(claves, reverse=True):
        inicio, fin = rango_ciclo(clave, movimientos)
        tiene_informacion = any(
            (fecha := fecha_movimiento(item.get("fecha", ""))) and inicio <= fecha <= fin
            for item in movimientos
        )
        etiqueta = etiqueta_ciclo(clave)
        if not tiene_informacion:
            etiqueta = f"{etiqueta} - Sin informacion"
        opciones.append(
            {
                "valor": clave,
                "etiqueta": etiqueta,
                "tiene_informacion": tiene_informacion,
            }
        )
    return opciones


def movimientos_de_ciclo(movimientos, clave):
    inicio, fin = rango_ciclo(clave, movimientos)
    items = [
        item
        for item in movimientos
        if (fecha := fecha_movimiento(item.get("fecha", ""))) and inicio <= fecha <= fin
    ]
    if not items:
        return [], inicio, fin
    saldo_anterior = (
        saldo_ciclo_anterior(movimientos, inicio)
        if ciclo_anterior_tiene_informacion(movimientos, inicio)
        else 0
    )
    if abs(saldo_anterior) >= 0.01:
        items.append(
            {
                "fecha": inicio.isoformat(),
                "tipo": "Ingreso",
                "categoria": "Saldo anterior",
                "descripcion": "Arrastre del ciclo anterior",
                "monto": saldo_anterior,
                "id": "",
            }
        )
    return sorted(items, key=lambda item: item["fecha"], reverse=True), inicio, fin


def asegurar_sueldo_automatico(hoy=None):
    hoy = hoy or fecha_hoy_chile()
    fecha_sueldo = fecha_sueldo_esperada(hoy)
    if not fecha_sueldo or hoy < fecha_sueldo:
        return False

    movimientos = leer_movimientos()
    if any(
        es_sueldo(item) and fecha_movimiento(item.get("fecha", "")) == fecha_sueldo
        for item in movimientos
    ):
        return False

    sueldo_anterior = ultimo_sueldo_antes_de(movimientos, fecha_sueldo)
    if not sueldo_anterior:
        return False

    guardar_movimiento(
        {
            "fecha": fecha_sueldo.isoformat(),
            "tipo": "Ingreso",
            "categoria": "Sueldo",
            "descripcion": "Sueldo",
            "monto": sueldo_anterior[1]["monto"],
        }
    )
    return True


def moneda(valor):
    return f"${valor:,.0f}".replace(",", ".")


app.jinja_env.filters["moneda"] = moneda
app.jinja_env.filters["fecha_corta"] = fecha_corta


def tooltip_operaciones(items, titulo):
    if not items:
        return f"{titulo}: {moneda(0)}"
    total = sum(item["monto"] for item in items)
    lineas = [f"{titulo}: {moneda(total)}"]
    for item in items[:6]:
        nombre = item.get("descripcion") or item.get("categoria") or item.get("persona") or "Sin detalle"
        lineas.append(f"{item.get('fecha', '')} · {nombre}: {moneda(item['monto'])}")
    if len(items) > 6:
        lineas.append(f"+{len(items) - 6} mas")
    return "\n".join(lineas)


app.jinja_env.filters["tooltip_operaciones"] = tooltip_operaciones


def operaciones_json(items, titulo):
    total = sum(item["monto"] for item in items)
    return json.dumps(
        {
            "titulo": titulo,
            "total": moneda(total),
            "items": [
                {
                    "fecha": item.get("fecha", ""),
                    "nombre": item.get("descripcion")
                    or item.get("categoria")
                    or item.get("persona")
                    or "Sin detalle",
                    "monto": moneda(item["monto"]),
                }
                for item in items
            ],
        },
        ensure_ascii=False,
    )


app.jinja_env.filters["operaciones_json"] = operaciones_json


def color_categoria(indice):
    colores = [
        "#1f6feb",
        "#c2410c",
        "#15803d",
        "#7c3aed",
        "#b45309",
        "#0f766e",
        "#be123c",
        "#4338ca",
    ]
    return colores[indice % len(colores)]


def preparar_segmentos(items):
    total = sum(item["monto"] for item in items)
    inicio = 0
    segmentos = []
    for indice, item in enumerate(items):
        porcentaje = item["monto"] / total * 100 if total else 0
        fin = inicio + porcentaje
        item["color"] = item.get("color") or color_categoria(indice)
        item["porcentaje"] = porcentaje
        item["offset"] = -inicio
        angulo = (inicio + porcentaje / 2) / 100 * 360 - 90
        radio = 37
        item["label_x"] = 60 + radio * math.cos(math.radians(angulo))
        item["label_y"] = 60 + radio * math.sin(math.radians(angulo))
        item["mostrar_label"] = porcentaje >= 7
        segmentos.append(f"{item['color']} {inicio:.2f}% {fin:.2f}%")
        inicio = fin
    return items, ", ".join(segmentos) if segmentos else "#e5e7eb 0% 100%", total


def calcular_semanas_restantes(hoy=None, fecha_fin=None):
    hoy = hoy or fecha_hoy_chile()
    if fecha_fin is None:
        ultimo_dia = calendar.monthrange(hoy.year, hoy.month)[1]
        fecha_fin = date(hoy.year, hoy.month, ultimo_dia)
    dias_restantes = max((fecha_fin - hoy).days + 1, 1)
    semanas = max(1, ceil(dias_restantes / 7))
    return semanas, dias_restantes, hoy


def rango_dashboard(hoy=None, movimientos=None):
    hoy = hoy or fecha_hoy_chile()
    movimientos = movimientos or []
    sueldo = ultimo_sueldo(movimientos)
    if sueldo:
        inicio = sueldo[0]
        proximo_mes = sumar_meses(inicio, 1)
        proximo_sueldo = fecha_movimiento(penultimo_dia_habil_mes(proximo_mes))
        fin = proximo_sueldo - timedelta(days=1)
        return inicio, fin

    inicio_mes = date(hoy.year, hoy.month, 1)
    inicio = inicio_mes - timedelta(days=7)
    ultimo_dia = calendar.monthrange(hoy.year, hoy.month)[1]
    fin = date(hoy.year, hoy.month, ultimo_dia)
    return inicio, fin


def fecha_movimiento(valor):
    try:
        return datetime.strptime(valor, "%Y-%m-%d").date()
    except ValueError:
        return None


def periodo_movimiento(valor):
    fecha = fecha_movimiento(valor)
    if not fecha:
        return ""
    return f"{fecha.year}-{fecha.month:02d}"


def descripcion_automatizacion(item):
    return item["descripcion"] or item["categoria"]


def mismo_texto(valor, otro):
    return (valor or "").strip().lower() == (otro or "").strip().lower()


def movimiento_coincide_automatizacion(movimiento, automatizacion):
    descripcion = automatizacion.get("descripcion", "").strip()
    descripcion_coincide = (
        not descripcion
        or mismo_texto(movimiento.get("descripcion"), descripcion_automatizacion(automatizacion))
    )
    return (
        movimiento.get("tipo") == automatizacion.get("tipo")
        and mismo_texto(movimiento.get("categoria"), automatizacion.get("categoria"))
        and descripcion_coincide
        and abs(float(movimiento.get("monto") or 0) - float(automatizacion.get("monto") or 0)) < 0.01
    )


def actualizar_automatizaciones_por_movimiento(movimiento, movimientos_actuales):
    periodo = periodo_movimiento_ciclo(movimiento.get("fecha", ""), movimientos_actuales)
    if not periodo:
        return

    automatizaciones = leer_automatizaciones()
    hubo_cambios = False
    for item in automatizaciones:
        if item.get("ultimo_confirmado") != periodo:
            continue
        if not movimiento_coincide_automatizacion(movimiento, item):
            continue
        existe_movimiento = any(
            periodo_movimiento_ciclo(actual.get("fecha", ""), movimientos_actuales) == periodo
            and movimiento_coincide_automatizacion(actual, item)
            for actual in movimientos_actuales
        )
        if not existe_movimiento:
            item["ultimo_confirmado"] = ""
            item["ticket_ultimo"] = ""
            hubo_cambios = True

    if hubo_cambios:
        escribir_automatizaciones(automatizaciones)


def sincronizar_automatizaciones_confirmadas():
    asegurar_sueldo_automatico()
    automatizaciones = leer_automatizaciones()
    movimientos = leer_movimientos()
    periodo = periodo_actual()
    hubo_cambios = False

    for item in automatizaciones:
        if esta_anulada(item, periodo):
            continue

        existe_en_periodo_actual = any(
            periodo_movimiento_ciclo(movimiento.get("fecha", ""), movimientos) == periodo
            and movimiento_coincide_automatizacion(movimiento, item)
            for movimiento in movimientos
        )
        if existe_en_periodo_actual and item.get("ultimo_confirmado") != periodo:
            item["ultimo_confirmado"] = periodo
            item["ticket_ultimo"] = ""
            hubo_cambios = True
            continue

        periodo_confirmado = item.get("ultimo_confirmado", "")
        if not periodo_confirmado:
            continue
        existe_movimiento = any(
            periodo_movimiento_ciclo(movimiento.get("fecha", ""), movimientos) == periodo_confirmado
            and movimiento_coincide_automatizacion(movimiento, item)
            for movimiento in movimientos
        )
        if not existe_movimiento:
            item["ultimo_confirmado"] = ""
            item["ticket_ultimo"] = ""
            hubo_cambios = True

    if hubo_cambios:
        escribir_automatizaciones(automatizaciones)

    return automatizaciones


def esta_anulada(item, periodo=None):
    periodo = periodo or periodo_actual()
    return item.get("ultimo_anulado") == periodo


def clave_periodo(fecha, vista):
    if vista == "diaria":
        return fecha.isoformat()
    if vista == "semanal":
        anio, semana, _ = fecha.isocalendar()
        return f"{anio}-S{semana:02d}"
    return f"{fecha.year}-{fecha.month:02d}"


def etiqueta_periodo(fecha, vista):
    if vista == "diaria":
        return fecha.strftime("%d-%m")
    if vista == "semanal":
        return f"S{fecha.isocalendar().week:02d}"
    return fecha.strftime("%m")


def rango_historico(vista, fechas):
    if vista == "diaria":
        fin = max(fechas)
        ventana_inicio = fin - timedelta(days=30)
        fechas_relevantes = [fecha for fecha in fechas if ventana_inicio <= fecha <= fin]
        inicio = min(fechas_relevantes) if fechas_relevantes else ventana_inicio
        return inicio, fin
    if vista == "semanal":
        fin = max(fechas)
        fin = fin - timedelta(days=fin.weekday())
        ventana_inicio = fin - timedelta(weeks=11)
        fechas_relevantes = [fecha for fecha in fechas if ventana_inicio <= fecha <= fin]
        primer_dato = min(fechas_relevantes) if fechas_relevantes else ventana_inicio
        inicio_datos = primer_dato - timedelta(days=primer_dato.weekday())
        inicio = max(inicio_datos, ventana_inicio)
        return inicio, fin
    fin = max(fechas)
    fin = date(fin.year, fin.month, 1)
    mes = fin.month - 11
    anio = fin.year
    while mes <= 0:
        mes += 12
        anio -= 1
    ventana_inicio = date(anio, mes, 1)
    fechas_relevantes = [fecha for fecha in fechas if ventana_inicio <= fecha <= fin]
    primer_dato = min(fechas_relevantes) if fechas_relevantes else ventana_inicio
    inicio_datos = date(primer_dato.year, primer_dato.month, 1)
    inicio = max(inicio_datos, ventana_inicio)
    return inicio, fin


def marcas_eje_y(maximo):
    if maximo <= 0:
        return [0]
    return [maximo * paso / 4 for paso in range(4, -1, -1)]


def construir_periodos(vista, fechas):
    periodos = {}
    if not fechas:
        return {}

    def nuevo_periodo(fecha):
        clave = clave_periodo(fecha, vista)
        periodos.setdefault(
            clave,
            {
                "periodo": clave,
                "etiqueta": etiqueta_periodo(fecha, vista),
                "ingresos": 0,
                "gastos": 0,
                "ahorros": 0,
                "me_deben": 0,
                "debo": 0,
                "balance": 0,
                "detalle_ingresos": [],
                "detalle_gastos": [],
                "detalle_ahorros": [],
                "detalle_me_deben": [],
                "detalle_debo": [],
            },
        )

    inicio, fin = rango_historico(vista, fechas)
    if vista == "diaria":
        actual = inicio
        while actual <= fin:
            nuevo_periodo(actual)
            actual += timedelta(days=1)
    elif vista == "semanal":
        actual = inicio - timedelta(days=inicio.weekday())
        while actual <= fin:
            nuevo_periodo(actual)
            actual += timedelta(days=7)
    else:
        actual = date(inicio.year, inicio.month, 1)
        while actual <= fin:
            nuevo_periodo(actual)
            if actual.month == 12:
                actual = date(actual.year + 1, 1, 1)
            else:
                actual = date(actual.year, actual.month + 1, 1)

    return periodos


def resumen_historico(vista):
    movimientos = leer_movimientos()
    deudas_lista = leer_deudas()
    fechas_movimientos = []
    fechas_deudas = []

    for item in movimientos:
        fecha = fecha_movimiento(item["fecha"])
        if fecha:
            fechas_movimientos.append(fecha)
    for item in deudas_lista:
        fecha = fecha_movimiento(item["fecha"])
        if fecha:
            fechas_deudas.append(fecha)

    periodos = construir_periodos(vista, fechas_movimientos)
    periodos_deudas = construir_periodos(vista, fechas_deudas)

    for item in movimientos:
        fecha = fecha_movimiento(item["fecha"])
        if not fecha:
            continue
        clave = clave_periodo(fecha, vista)
        if clave not in periodos:
            continue
        periodo = periodos[clave]
        if item["tipo"] == "Ingreso":
            periodo["ingresos"] += item["monto"]
            periodo["detalle_ingresos"].append(item)
        elif item["tipo"] == "Gasto":
            periodo["gastos"] += item["monto"]
            periodo["detalle_gastos"].append(item)
        elif item["tipo"] == "Ahorro":
            periodo["ahorros"] += item["monto"]
            periodo["detalle_ahorros"].append(item)

    for item in deudas_lista:
        fecha = fecha_movimiento(item["fecha"])
        if not fecha:
            continue
        clave = clave_periodo(fecha, vista)
        if clave not in periodos_deudas:
            continue
        periodo = periodos_deudas[clave]
        if item["tipo"] == "Me deben":
            periodo["me_deben"] += item["monto"]
            periodo["detalle_me_deben"].append(item)
        elif item["tipo"] == "Debo":
            periodo["debo"] += item["monto"]
            periodo["detalle_debo"].append(item)

    filas = [periodos[clave] for clave in sorted(periodos)]
    filas_deudas = [periodos_deudas[clave] for clave in sorted(periodos_deudas)]
    for fila in filas:
        fila["balance"] = fila["ingresos"] - fila["gastos"] - fila["ahorros"]
    for fila in filas_deudas:
        fila["balance"] = fila["me_deben"] - fila["debo"]

    maximo = max(
        [
            valor
            for fila in filas
            for valor in [
                fila["ingresos"],
                fila["gastos"],
                fila["ahorros"],
                fila["me_deben"],
                fila["debo"],
            ]
        ]
        or [1]
    )
    maximo_deudas = max(
        [valor for fila in filas_deudas for valor in [fila["me_deben"], fila["debo"]]]
        or [1]
    )
    return filas, maximo, marcas_eje_y(maximo), filas_deudas, maximo_deudas, marcas_eje_y(maximo_deudas)


def calcular_dashboard(filtrar_periodo=False):
    hoy = fecha_hoy_chile()
    asegurar_sueldo_automatico(hoy)
    movimientos = leer_movimientos()
    periodo_inicio = None
    periodo_fin = None
    saldo_anterior = 0
    if filtrar_periodo:
        periodo_inicio, periodo_fin = rango_dashboard(hoy, movimientos)
        saldo_anterior = (
            saldo_ciclo_anterior(movimientos, periodo_inicio)
            if ciclo_anterior_tiene_informacion(movimientos, periodo_inicio)
            else 0
        )
        movimientos = [
            item
            for item in movimientos
            if (fecha := fecha_movimiento(item["fecha"]))
            and periodo_inicio <= fecha <= periodo_fin
        ]
        if abs(saldo_anterior) >= 0.01:
            movimientos.append(
                {
                    "fecha": periodo_inicio.isoformat(),
                    "tipo": "Ingreso",
                    "categoria": "Saldo anterior",
                    "descripcion": "Arrastre del ciclo anterior",
                    "monto": saldo_anterior,
                    "id": "",
                }
            )
    semanas, dias_restantes, hoy = calcular_semanas_restantes(hoy, periodo_fin)
    movimientos = sorted(movimientos, key=lambda item: item["fecha"], reverse=True)
    ingresos_lista = [item for item in movimientos if item["tipo"] == "Ingreso"]
    deudas_lista = [
        item
        for item in movimientos
        if item["tipo"] == "Gasto" and item.get("categoria", "").strip().lower() == "deuda"
    ]
    gastos_lista = [
        item
        for item in movimientos
        if item["tipo"] == "Gasto" and item not in deudas_lista
    ]
    ahorros_lista = [item for item in movimientos if item["tipo"] == "Ahorro"]
    ingresos = sum(item["monto"] for item in ingresos_lista)
    gastos = sum(item["monto"] for item in gastos_lista)
    deuda = sum(item["monto"] for item in deudas_lista)
    ahorros = sum(item["monto"] for item in ahorros_lista)
    disponible = ingresos - gastos - deuda - ahorros
    cuota_semanal = disponible / semanas if semanas > 0 else 0
    return {
        "movimientos": movimientos,
        "ingresos_lista": ingresos_lista,
        "gastos_lista": gastos_lista,
        "deudas_lista": deudas_lista,
        "ahorros_lista": ahorros_lista,
        "ingresos": ingresos,
        "gastos": gastos,
        "deuda": deuda,
        "ahorros": ahorros,
        "balance": ingresos - gastos - deuda,
        "disponible": disponible,
        "semanas": semanas,
        "dias_restantes": dias_restantes,
        "fecha_calculo": hoy,
        "periodo_inicio": periodo_inicio,
        "periodo_fin": periodo_fin,
        "cuota_semanal": cuota_semanal,
    }


def calcular_planificacion():
    datos = calcular_dashboard(filtrar_periodo=True)
    periodo_inicio = datos["periodo_inicio"]
    periodo_fin = datos["periodo_fin"]
    periodo = periodo_actual()
    movimientos = datos["movimientos"]

    automatizaciones = sincronizar_automatizaciones_confirmadas()
    fijos_pendientes = []
    for item in automatizaciones:
        if not item["activo"] or esta_anulada(item, periodo):
            continue
        if item.get("ultimo_confirmado") == periodo:
            continue
        existe_movimiento = any(
            (fecha := fecha_movimiento(movimiento.get("fecha", "")))
            and periodo_inicio <= fecha <= periodo_fin
            and movimiento_coincide_automatizacion(movimiento, item)
            for movimiento in movimientos
        )
        if not existe_movimiento:
            fijos_pendientes.append(item)

    planificaciones = []
    for item in leer_planificaciones():
        fecha = fecha_movimiento(item.get("fecha", ""))
        if fecha and periodo_inicio <= fecha <= periodo_fin:
            planificaciones.append(item)

    fijos_gastos = sum(item["monto"] for item in fijos_pendientes if item["tipo"] == "Gasto")
    fijos_ahorros = sum(item["monto"] for item in fijos_pendientes if item["tipo"] == "Ahorro")
    deudas_cuotas = [
        item
        for item in leer_deudas()
        if item.get("tipo") == "Debo"
        and item.get("modalidad") == "Cuotas"
        and item.get("estado") != "Pagada"
    ]
    cuotas_mes = sum(monto_cuota(item) for item in deudas_cuotas)
    fijos_gastos += cuotas_mes
    gastos_fijos_ciclo = sum(
        item["monto"]
        for item in automatizaciones
        if item.get("activo") and item.get("tipo") == "Gasto" and not esta_anulada(item, periodo)
    ) + cuotas_mes
    plan_ingresos = sum(item["monto"] for item in planificaciones if item["tipo"] == "Ingreso")
    plan_gastos = sum(item["monto"] for item in planificaciones if item["tipo"] == "Gasto")
    plan_ahorros = sum(item["monto"] for item in planificaciones if item["tipo"] == "Ahorro")

    ingresos = datos["ingresos"] + plan_ingresos
    gastos = datos["gastos"] + datos["deuda"] + fijos_gastos + plan_gastos
    ahorros = datos["ahorros"] + fijos_ahorros + plan_ahorros
    disponible = ingresos - gastos - ahorros

    asignaciones = asignaciones_macro_dict()
    variables = {macro: 0.0 for macro in MACROS_ASIGNABLES}
    for item in movimientos + planificaciones:
        tipo = item.get("tipo")
        if tipo not in {"Gasto", "Ahorro"}:
            continue
        categoria = (item.get("categoria") or "").strip().lower()
        es_fijo = tipo == "Gasto" and any(
            automatizacion.get("activo")
            and movimiento_coincide_automatizacion(item, automatizacion)
            for automatizacion in automatizaciones
        )
        if es_fijo:
            continue
        macro = asignaciones.get((tipo, categoria))
        if macro:
            variables[macro] += float(item.get("monto") or 0)

    valores_macro = {
        "Gastos fijos": gastos_fijos_ciclo,
        **variables,
    }
    valores_macro["Libre"] = max(disponible, 0)
    metas = leer_metas()
    for meta in metas:
        meta["objetivo"] = ingresos * meta["porcentaje"] / 100
        meta["actual"] = valores_macro.get(meta["macro"], 0)
        objetivo = meta["objetivo"]
        meta["avance"] = min(meta["actual"] / objetivo * 100, 100) if objetivo > 0 else 0
        meta["excedida"] = objetivo > 0 and meta["actual"] > objetivo

    return {
        "periodo_inicio": periodo_inicio,
        "periodo_fin": periodo_fin,
        "ingresos_base": datos["ingresos"],
        "gastos_base": datos["gastos"] + datos["deuda"],
        "ahorros_base": datos["ahorros"],
        "fijos_pendientes": fijos_pendientes,
        "fijos_gastos": fijos_gastos,
        "fijos_ahorros": fijos_ahorros,
        "deudas_cuotas": deudas_cuotas,
        "cuotas_mes": cuotas_mes,
        "gastos_fijos_ciclo": gastos_fijos_ciclo,
        "planificaciones": planificaciones,
        "plan_ingresos": plan_ingresos,
        "plan_gastos": plan_gastos,
        "plan_ahorros": plan_ahorros,
        "ingresos": ingresos,
        "gastos": gastos,
        "ahorros": ahorros,
        "disponible": disponible,
        "metas": metas,
        "metas_total": sum(item["porcentaje"] for item in metas),
    }


@app.route("/")
def index():
    return render_template("index.html", **calcular_dashboard(filtrar_periodo=True))


@app.route("/planificacion", methods=["GET", "POST"])
def planificacion():
    if request.method == "POST":
        if request.form.get("accion") == "guardar_metas":
            metas = []
            for macro, predeterminado in METAS_PREDEFINIDAS:
                try:
                    porcentaje = float(request.form.get(f"meta_{macro}", predeterminado) or 0)
                except ValueError:
                    porcentaje = predeterminado
                metas.append({"macro": macro, "porcentaje": min(max(porcentaje, 0), 100)})
            escribir_metas(metas)
            return redirect(url_for("planificacion", _anchor="metas"))
        tipo = request.form.get("tipo", "Gasto")
        if tipo not in TIPOS_VALIDOS:
            tipo = "Gasto"
        guardar_planificacion(
            {
                "fecha": request.form.get("fecha", ""),
                "tipo": tipo,
                "categoria": request.form.get("categoria", "").strip(),
                "descripcion": request.form.get("descripcion", "").strip(),
                "monto": float(request.form.get("monto", 0) or 0),
            }
        )
        return redirect(url_for("planificacion"))

    return render_template("planificacion.html", **calcular_planificacion())


def categorias_para_macros():
    categorias = {}

    def agregar(tipo, categoria):
        categoria = (categoria or "").strip()
        if tipo not in {"Gasto", "Ahorro"} or not categoria:
            return
        categorias.setdefault((tipo, categoria.lower()), categoria)

    for item in leer_categorias():
        agregar(item.get("tipo", ""), item.get("categoria", ""))
    for coleccion in (leer_movimientos(), leer_automatizaciones(), leer_planificaciones()):
        for item in coleccion:
            agregar(item.get("tipo", ""), item.get("categoria", ""))

    guardadas = asignaciones_macro_dict()
    return [
        {
            "tipo": tipo,
            "categoria": categoria,
            "macro": guardadas.get((tipo, clave), ""),
        }
        for (tipo, clave), categoria in sorted(
            categorias.items(), key=lambda item: (item[0][0], item[1].lower())
        )
    ]


def categorias_administrables():
    categorias = {}
    for item in leer_categorias():
        tipo = item.get("tipo", "")
        categoria = item.get("categoria", "").strip()
        if tipo in TIPOS_VALIDOS and categoria:
            categorias[(tipo, categoria.lower())] = {"tipo": tipo, "categoria": categoria}
    for coleccion in (leer_movimientos(), leer_automatizaciones(), leer_planificaciones()):
        for item in coleccion:
            tipo = item.get("tipo", "")
            categoria = item.get("categoria", "").strip()
            if tipo in TIPOS_VALIDOS and categoria:
                categorias.setdefault(
                    (tipo, categoria.lower()), {"tipo": tipo, "categoria": categoria}
                )
    return sorted(categorias.values(), key=lambda item: (item["tipo"], item["categoria"].lower()))


def reemplazar_categoria_eliminada(tipo, categoria):
    categoria_clave = categoria.strip().lower()
    colecciones = [
        (leer_movimientos, escribir_movimientos),
        (leer_automatizaciones, escribir_automatizaciones),
        (leer_planificaciones, escribir_planificaciones),
    ]
    for leer, escribir in colecciones:
        items = leer()
        hubo_cambios = False
        for item in items:
            if (
                item.get("tipo") == tipo
                and item.get("categoria", "").strip().lower() == categoria_clave
            ):
                item["categoria"] = CATEGORIA_SIN_ASIGNAR
                hubo_cambios = True
        if hubo_cambios:
            escribir(items)


@app.route("/macrocategorias/categorias/agregar", methods=["POST"])
def agregar_categoria():
    tipo = request.form.get("tipo", "")
    categoria = request.form.get("categoria", "").strip()
    if tipo in TIPOS_VALIDOS and categoria:
        categorias = leer_categorias()
        if not any(
            item.get("tipo") == tipo
            and item.get("categoria", "").strip().lower() == categoria.lower()
            for item in categorias
        ):
            categorias.append({"tipo": tipo, "categoria": categoria})
            escribir_categorias(categorias)
    return redirect(url_for("macrocategorias", _anchor="categorias"))


@app.route("/macrocategorias/categorias/eliminar", methods=["POST"])
def eliminar_categoria():
    tipo = request.form.get("tipo", "")
    categoria = request.form.get("categoria", "").strip()
    if (
        tipo in TIPOS_VALIDOS
        and categoria
        and categoria.lower() != CATEGORIA_SIN_ASIGNAR.lower()
    ):
        reemplazar_categoria_eliminada(tipo, categoria)
        escribir_categorias(
            [
                item
                for item in leer_categorias()
                if not (
                    item.get("tipo") == tipo
                    and item.get("categoria", "").strip().lower() == categoria.lower()
                )
            ]
        )
        escribir_macro_categorias(
            [
                item
                for item in leer_macro_categorias()
                if not (
                    item.get("tipo") == tipo
                    and item.get("categoria", "").strip().lower() == categoria.lower()
                )
            ]
        )
    return redirect(url_for("macrocategorias", _anchor="categorias"))


@app.route("/macrocategorias", methods=["GET", "POST"])
def macrocategorias():
    categorias = categorias_para_macros()
    if request.method == "POST":
        asignaciones = []
        for indice, item in enumerate(categorias):
            macro = request.form.get(f"macro_{indice}", "")
            if macro in MACROS_ASIGNABLES:
                asignaciones.append({**item, "macro": macro})
        escribir_macro_categorias(asignaciones)
        return redirect(url_for("macrocategorias"))

    gastos_fijos = [
        item
        for item in leer_automatizaciones()
        if item.get("activo") and item.get("tipo") == "Gasto"
    ]
    return render_template(
        "macrocategorias.html",
        categorias=categorias,
        categorias_administrables=categorias_administrables(),
        macros=MACROS_ASIGNABLES,
        gastos_fijos=gastos_fijos,
    )


@app.route("/planificacion/eliminar/<int:planificacion_id>", methods=["POST"])
def eliminar_planificacion(planificacion_id):
    planificaciones = leer_planificaciones()
    if 0 <= planificacion_id < len(planificaciones):
        planificaciones.pop(planificacion_id)
        escribir_planificaciones(planificaciones)
    return redirect(url_for("planificacion"))


@app.route("/agregar", methods=["GET", "POST"])
def agregar():
    if request.method == "POST":
        movimiento = {
            "fecha": request.form.get("fecha", ""),
            "tipo": request.form.get("tipo", "Gasto")
            if request.form.get("tipo", "Gasto") in TIPOS_VALIDOS
            else "Gasto",
            "categoria": request.form.get("categoria", "").strip(),
            "descripcion": request.form.get("descripcion", "").strip(),
            "monto": float(request.form.get("monto", 0) or 0),
        }
        guardar_movimiento(movimiento)
        return redirect(url_for("index"))

    return render_template("agregar.html")


@app.route("/agregar/<tipo>", methods=["GET", "POST"])
def agregar_por_tipo(tipo):
    tipos = {
        "ingreso": ("Ingreso", "Agregar ingreso", "Sueldo, venta, bono..."),
        "gasto": ("Gasto", "Agregar gasto", "Comida, transporte, arriendo..."),
        "ahorro": ("Ahorro", "Agregar ahorro", "Emergencia, viaje, inversion..."),
    }
    if tipo not in tipos:
        return redirect(url_for("agregar"))

    tipo_movimiento, titulo, ayuda_categoria = tipos[tipo]
    if request.method == "POST":
        movimiento = {
            "fecha": request.form.get("fecha", ""),
            "tipo": tipo_movimiento,
            "categoria": request.form.get("categoria", "").strip(),
            "descripcion": request.form.get("descripcion", "").strip(),
            "monto": float(request.form.get("monto", 0) or 0),
        }
        guardar_movimiento(movimiento)
        return redirect(url_for("index"))

    return render_template(
        "agregar.html",
        tipo=tipo_movimiento,
        titulo=titulo,
        ayuda_categoria=ayuda_categoria,
        sueldo_fecha=penultimo_dia_habil_mes() if tipo_movimiento == "Ingreso" else "",
    )


@app.route("/agregar/sueldo", methods=["POST"])
def agregar_sueldo():
    guardar_movimiento(
        {
            "fecha": penultimo_dia_habil_mes(),
            "tipo": "Ingreso",
            "categoria": "Sueldo",
            "descripcion": "Sueldo",
            "monto": float(request.form.get("monto", 0) or 0),
        }
    )
    return redirect(url_for("index"))


@app.route("/editar/<int:movimiento_id>", methods=["GET", "POST"])
def editar(movimiento_id):
    volver_ciclo = request.values.get("volver_ciclo", "").strip()
    volver_busqueda = request.values.get("volver_q", "").strip()
    volver_tipo = request.values.get("volver_tipo", "Todos")
    parametros_resumen = {}
    if volver_ciclo:
        parametros_resumen["ciclo"] = volver_ciclo
    if volver_busqueda:
        parametros_resumen["q"] = volver_busqueda
    if volver_tipo in TIPOS_VALIDOS:
        parametros_resumen["tipo"] = volver_tipo
    volver_a = f"{url_for('resumen', **parametros_resumen)}#movimientos"

    movimientos = leer_movimientos()
    if movimiento_id < 0 or movimiento_id >= len(movimientos):
        return redirect(volver_a)

    movimiento = movimientos[movimiento_id]
    if request.method == "POST":
        tipo = request.form.get("tipo", movimiento["tipo"])
        movimiento_anterior = movimiento.copy()
        movimientos[movimiento_id] = {
            "fecha": request.form.get("fecha", ""),
            "tipo": tipo if tipo in TIPOS_VALIDOS else movimiento["tipo"],
            "categoria": request.form.get("categoria", "").strip(),
            "descripcion": request.form.get("descripcion", "").strip(),
            "monto": float(request.form.get("monto", 0) or 0),
        }
        escribir_movimientos(movimientos)
        actualizar_automatizaciones_por_movimiento(movimiento_anterior, movimientos)
        return redirect(volver_a)

    return render_template(
        "agregar.html",
        titulo="Editar movimiento",
        tipo=None,
        movimiento=movimiento,
        ayuda_categoria="Categoria del movimiento",
        volver_a=volver_a,
        volver_ciclo=volver_ciclo,
        volver_busqueda=volver_busqueda,
        volver_tipo=volver_tipo,
    )


@app.route("/eliminar/<int:movimiento_id>", methods=["POST"])
def eliminar(movimiento_id):
    volver_ciclo = request.form.get("volver_ciclo", "").strip()
    volver_busqueda = request.form.get("volver_q", "").strip()
    volver_tipo = request.form.get("volver_tipo", "Todos")
    parametros_resumen = {}
    if volver_ciclo:
        parametros_resumen["ciclo"] = volver_ciclo
    if volver_busqueda:
        parametros_resumen["q"] = volver_busqueda
    if volver_tipo in TIPOS_VALIDOS:
        parametros_resumen["tipo"] = volver_tipo

    movimientos = leer_movimientos()
    if 0 <= movimiento_id < len(movimientos):
        movimiento_eliminado = movimientos.pop(movimiento_id)
        escribir_movimientos(movimientos)
        actualizar_automatizaciones_por_movimiento(movimiento_eliminado, movimientos)
    return redirect(f"{url_for('resumen', **parametros_resumen)}#movimientos")


@app.route("/resumen/agregar", methods=["POST"])
def agregar_desde_resumen():
    tipo = request.form.get("tipo", "Gasto")
    if tipo not in TIPOS_VALIDOS:
        tipo = "Gasto"
    guardar_movimiento(
        {
            "fecha": request.form.get("fecha", ""),
            "tipo": tipo,
            "categoria": request.form.get("categoria", "").strip(),
            "descripcion": request.form.get("descripcion", "").strip(),
            "monto": float(request.form.get("monto", 0) or 0),
        }
    )
    return redirect(f"{url_for('resumen')}#movimientos")


@app.route("/gastos-fijos", methods=["GET", "POST"])
@app.route("/automatizacion", methods=["GET", "POST"])
def automatizacion():
    if request.method == "POST":
        tipo = request.form.get("tipo", "Gasto")
        if tipo not in TIPOS_AUTOMATIZACION:
            tipo = "Gasto"
        try:
            dia_mes = int(request.form.get("dia_mes", 1) or 1)
        except ValueError:
            dia_mes = 1
        automatizaciones = leer_automatizaciones()
        automatizaciones.append(
            {
                "tipo": tipo,
                "categoria": request.form.get("categoria", "").strip(),
                "descripcion": request.form.get("descripcion", "").strip(),
                "monto": float(request.form.get("monto", 0) or 0),
                "dia_mes": min(max(dia_mes, 1), 31),
                "activo": True,
                "ultimo_confirmado": "",
                "ticket_ultimo": "",
                "ultimo_anulado": "",
                "razon_anulado": "",
            }
        )
        escribir_automatizaciones(automatizaciones)
        return redirect(url_for("automatizacion"))

    asegurar_sueldo_automatico()
    periodo = periodo_actual()
    automatizaciones = sorted(
        sincronizar_automatizaciones_confirmadas(),
        key=lambda item: (
            item["categoria"].lower(),
            item["descripcion"].lower(),
            item["tipo"].lower(),
        ),
    )
    pendientes = [
        item
        for item in automatizaciones
        if item["activo"]
        and item.get("ultimo_confirmado") != periodo
        and not esta_anulada(item, periodo)
    ]
    confirmadas = [
        item
        for item in automatizaciones
        if item["activo"] and item.get("ultimo_confirmado") == periodo
    ]
    anuladas = [
        item
        for item in automatizaciones
        if item["activo"] and esta_anulada(item, periodo)
    ]
    total_gastos_fijos = sum(
        item["monto"]
        for item in automatizaciones
        if item["activo"] and item["tipo"] == "Gasto" and not esta_anulada(item, periodo)
    )
    total_gastos_fijos_sin_arriendo = sum(
        item["monto"]
        for item in automatizaciones
        if item["activo"]
        and item["tipo"] == "Gasto"
        and not esta_anulada(item, periodo)
        and "arriendo" not in f"{item['categoria']} {item['descripcion']}".lower()
    )
    total_ahorros_planificados = sum(
        item["monto"]
        for item in automatizaciones
        if item["activo"] and item["tipo"] == "Ahorro" and not esta_anulada(item, periodo)
    )
    compromiso_mensual = total_gastos_fijos + total_ahorros_planificados
    sueldo = ultimo_sueldo(leer_movimientos())
    sueldo_monto = sueldo[1]["monto"] if sueldo else 0
    return render_template(
        "automatizacion.html",
        automatizaciones=automatizaciones,
        pendientes=pendientes,
        confirmadas=confirmadas,
        anuladas=anuladas,
        periodo=periodo,
        total_gastos_fijos=total_gastos_fijos,
        total_gastos_fijos_sin_arriendo=total_gastos_fijos_sin_arriendo,
        total_ahorros_planificados=total_ahorros_planificados,
        compromiso_mensual=compromiso_mensual,
        sueldo_menos_compromisos=sueldo_monto - compromiso_mensual,
    )


@app.route("/automatizacion/confirmar/<int:automatizacion_id>", methods=["POST"])
def confirmar_automatizacion(automatizacion_id):
    automatizaciones = leer_automatizaciones()
    if automatizacion_id < 0 or automatizacion_id >= len(automatizaciones):
        return redirect(url_for("automatizacion"))

    item = automatizaciones[automatizacion_id]
    descripcion = item["descripcion"] or item["categoria"]
    fecha = request.form.get("fecha") or proxima_fecha_mensual(item["dia_mes"])
    guardar_movimiento(
        {
            "fecha": fecha,
            "tipo": item["tipo"],
            "categoria": item["categoria"],
            "descripcion": descripcion,
            "monto": item["monto"],
        }
    )
    item["ultimo_confirmado"] = periodo_movimiento_ciclo(fecha) or periodo_actual()
    item["ticket_ultimo"] = ""
    item["ultimo_anulado"] = ""
    item["razon_anulado"] = ""
    automatizaciones[automatizacion_id] = item
    escribir_automatizaciones(automatizaciones)
    return redirect(url_for("automatizacion"))


@app.route("/automatizacion/anular/<int:automatizacion_id>", methods=["POST"])
def anular_automatizacion(automatizacion_id):
    automatizaciones = leer_automatizaciones()
    if automatizacion_id < 0 or automatizacion_id >= len(automatizaciones):
        return redirect(url_for("automatizacion"))

    razon = request.form.get("razon_anulado", "").strip()
    if not razon:
        return redirect(url_for("automatizacion"))

    item = automatizaciones[automatizacion_id]
    item["ultimo_anulado"] = periodo_actual()
    item["razon_anulado"] = razon
    item["ultimo_confirmado"] = ""
    item["ticket_ultimo"] = ""
    automatizaciones[automatizacion_id] = item
    escribir_automatizaciones(automatizaciones)
    return redirect(url_for("automatizacion"))


@app.route("/automatizacion/desanular/<int:automatizacion_id>", methods=["POST"])
def desanular_automatizacion(automatizacion_id):
    automatizaciones = leer_automatizaciones()
    if automatizacion_id < 0 or automatizacion_id >= len(automatizaciones):
        return redirect(url_for("automatizacion"))

    item = automatizaciones[automatizacion_id]
    if esta_anulada(item):
        item["ultimo_anulado"] = ""
        item["razon_anulado"] = ""
        automatizaciones[automatizacion_id] = item
        escribir_automatizaciones(automatizaciones)
    return redirect(url_for("automatizacion"))


@app.route("/gastos-fijos/descargar")
def descargar_gastos_fijos():
    periodo = periodo_actual()
    salida = []
    salida.append(
        [
            "Periodo",
            "Tipo",
            "Descripcion",
            "Categoria",
            "Dia del mes",
            "Monto",
            "Estado",
        ]
    )
    for item in sorted(
        sincronizar_automatizaciones_confirmadas(),
        key=lambda fila: (
            fila["descripcion"].lower(),
            fila["categoria"].lower(),
            fila["tipo"].lower(),
        ),
    ):
        if item.get("ultimo_confirmado") == periodo:
            estado = "Confirmado"
        elif esta_anulada(item, periodo):
            estado = f"Anulado: {item.get('razon_anulado', '')}"
        else:
            estado = "Pendiente"
        salida.append(
            [
                periodo,
                item["tipo"],
                item["descripcion"] or "Sin descripcion",
                item["categoria"] or "Sin categoria",
                item["dia_mes"],
                int(item["monto"]) if item["monto"].is_integer() else item["monto"],
                estado,
            ]
        )

    contenido = "\ufeff" + "\n".join(
        ";".join(str(valor).replace(";", ",") for valor in fila) for fila in salida
    )
    return Response(
        contenido,
        mimetype="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f"attachment; filename=gastos_fijos_{periodo}.csv"
        },
    )


@app.route("/automatizacion/editar/<int:automatizacion_id>", methods=["GET", "POST"])
def editar_automatizacion(automatizacion_id):
    automatizaciones = leer_automatizaciones()
    if automatizacion_id < 0 or automatizacion_id >= len(automatizaciones):
        return redirect(url_for("automatizacion"))

    item = automatizaciones[automatizacion_id]
    if request.method == "POST":
        tipo = request.form.get("tipo", item["tipo"])
        if tipo not in TIPOS_AUTOMATIZACION:
            tipo = item["tipo"]
        try:
            dia_mes = int(request.form.get("dia_mes", item["dia_mes"]) or item["dia_mes"])
        except ValueError:
            dia_mes = item["dia_mes"]
        item.update(
            {
                "tipo": tipo,
                "categoria": request.form.get("categoria", "").strip(),
                "descripcion": request.form.get("descripcion", "").strip(),
                "monto": float(request.form.get("monto", 0) or 0),
                "dia_mes": min(max(dia_mes, 1), 31),
            }
        )
        automatizaciones[automatizacion_id] = item
        escribir_automatizaciones(automatizaciones)
        return redirect(url_for("automatizacion"))

    return render_template("editar_automatizacion.html", item=item)


@app.route("/automatizacion/eliminar/<int:automatizacion_id>", methods=["POST"])
def eliminar_automatizacion(automatizacion_id):
    automatizaciones = leer_automatizaciones()
    if 0 <= automatizacion_id < len(automatizaciones):
        automatizaciones.pop(automatizacion_id)
        escribir_automatizaciones(automatizaciones)
    return redirect(url_for("automatizacion"))


@app.route("/deudas", methods=["GET", "POST"])
def deudas():
    if request.method == "POST":
        tipo = request.form.get("tipo", "Me deben")
        if tipo not in TIPOS_DEUDA:
            tipo = "Me deben"
        modalidad = request.form.get("modalidad", "Simple") if tipo == "Debo" else "Simple"
        if modalidad not in {"Simple", "Cuotas", "Prestamo"}:
            modalidad = "Simple"
        try:
            cuotas_total = max(int(request.form.get("cuotas_total", 1) or 1), 1)
        except ValueError:
            cuotas_total = 1
        ticket_deuda = uuid.uuid4().hex
        item = {
            "fecha": request.form.get("fecha", ""),
            "tipo": tipo,
            "persona": request.form.get("persona", "").strip(),
            "categoria": request.form.get("categoria", "").strip(),
            "descripcion": request.form.get("descripcion", "").strip(),
            "monto": float(request.form.get("monto", 0) or 0),
            "estado": "Pendiente",
            "fecha_pago": "",
            "modalidad": modalidad,
            "cuotas_total": cuotas_total if modalidad == "Cuotas" else 1,
            "cuotas_pagadas": 0,
            "ticket_deuda": ticket_deuda,
        }
        deudas_lista = leer_deudas()
        deudas_lista.append(item)
        escribir_deudas(deudas_lista)
        tipo_origen = None
        if tipo == "Me deben":
            tipo_origen = "Gasto"
        elif tipo == "Debo" and modalidad == "Prestamo":
            tipo_origen = "Ingreso"
        if tipo_origen:
            guardar_movimiento(
                {
                    "fecha": item["fecha"],
                    "tipo": tipo_origen,
                    "categoria": "Deuda",
                    "descripcion": item["descripcion"] or item["persona"],
                    "monto": item["monto"],
                    "ticket_deuda": ticket_deuda,
                    "origen_deuda": "origen",
                }
            )
        return redirect(url_for("deudas"))

    deudas_lista = leer_deudas()
    pendientes = [item for item in deudas_lista if item["estado"] != "Pagada"]
    pagadas = [item for item in deudas_lista if item["estado"] == "Pagada"]
    me_deben = [item for item in pendientes if item["tipo"] == "Me deben"]
    debo = [item for item in pendientes if item["tipo"] == "Debo"]
    total_me_deben = sum(item["monto"] for item in me_deben)
    total_debo = sum(item["monto"] for item in debo)
    return render_template(
        "deudas.html",
        deudas=deudas_lista,
        pendientes=pendientes,
        pagadas=pagadas,
        me_deben=me_deben,
        debo=debo,
        total_me_deben=total_me_deben,
        total_debo=total_debo,
        balance_deudas=total_me_deben - total_debo,
    )


@app.route("/deudas/pagar/<int:deuda_id>", methods=["POST"])
def pagar_deuda(deuda_id):
    deudas_lista = leer_deudas()
    if deuda_id < 0 or deuda_id >= len(deudas_lista):
        return redirect(url_for("deudas"))

    item = deudas_lista[deuda_id]
    if item["estado"] != "Pagada":
        fecha_pago = request.form.get("fecha_pago") or fecha_hoy_chile().isoformat()
        movimientos = leer_movimientos()
        mismo_ciclo = not pago_deuda_genera_movimiento(item, fecha_pago, movimientos)
        if mismo_ciclo:
            eliminar_movimientos_deuda(item.get("ticket_deuda", ""), "origen")
        else:
            tipo_movimiento = "Ingreso" if item["tipo"] == "Me deben" else "Gasto"
            monto_pago = monto_cuota(item) if item.get("modalidad") == "Cuotas" else item["monto"]
            guardar_movimiento(
                {
                    "fecha": fecha_pago,
                    "tipo": tipo_movimiento,
                    "categoria": item["categoria"] or "Deudas",
                    "descripcion": f"{item['tipo']} - {item['persona']} | {item['descripcion']}",
                    "monto": monto_pago,
                    "ticket_deuda": item.get("ticket_deuda", ""),
                    "origen_deuda": "pago",
                }
            )
        if item.get("modalidad") == "Cuotas":
            item["cuotas_pagadas"] = min(
                int(item.get("cuotas_pagadas") or 0) + 1,
                int(item.get("cuotas_total") or 1),
            )
            item["estado"] = (
                "Pagada" if item["cuotas_pagadas"] >= item["cuotas_total"] else "Pendiente"
            )
        else:
            item["estado"] = "Pagada"
        item["fecha_pago"] = fecha_pago
        deudas_lista[deuda_id] = item
        escribir_deudas(deudas_lista)
    return redirect(url_for("deudas"))


@app.route("/deudas/eliminar/<int:deuda_id>", methods=["POST"])
def eliminar_deuda(deuda_id):
    deudas_lista = leer_deudas()
    if 0 <= deuda_id < len(deudas_lista):
        item = deudas_lista.pop(deuda_id)
        eliminar_movimientos_deuda(item.get("ticket_deuda", ""))
        escribir_deudas(deudas_lista)
    return redirect(url_for("deudas"))


@app.route("/historico")
def historico():
    vista = request.args.get("vista", "diaria")
    if vista not in {"diaria", "semanal", "mensual"}:
        vista = "diaria"
    filtro_tipo = request.args.get("tipo", "Todos")
    if filtro_tipo not in TIPOS_VALIDOS | {"Todos"}:
        filtro_tipo = "Todos"
    filas, maximo, eje_y, filas_deudas, maximo_deudas, eje_y_deudas = resumen_historico(vista)
    totales = {
        "ingresos": sum(fila["ingresos"] for fila in filas),
        "gastos": sum(fila["gastos"] for fila in filas),
        "ahorros": sum(fila["ahorros"] for fila in filas),
        "me_deben": sum(fila["me_deben"] for fila in filas),
        "debo": sum(fila["debo"] for fila in filas),
    }
    datasets_por_tipo = {
        "Ingreso": {
            "label": "Ingresos",
            "values": [fila["ingresos"] for fila in filas],
            "color": "#15803d",
            "details": [operaciones_json(fila["detalle_ingresos"], "Ingresos") for fila in filas],
        },
        "Gasto": {
            "label": "Gastos",
            "values": [fila["gastos"] for fila in filas],
            "color": "#c2410c",
            "details": [operaciones_json(fila["detalle_gastos"], "Gastos") for fila in filas],
        },
        "Ahorro": {
            "label": "Ahorros",
            "values": [fila["ahorros"] for fila in filas],
            "color": "#7c3aed",
            "details": [operaciones_json(fila["detalle_ahorros"], "Ahorros") for fila in filas],
        },
    }
    datasets_movimientos = (
        list(datasets_por_tipo.values())
        if filtro_tipo == "Todos"
        else [datasets_por_tipo[filtro_tipo]]
    )
    return render_template(
        "historico.html",
        vista=vista,
        filtro_tipo=filtro_tipo,
        filas=filas,
        datasets_movimientos=datasets_movimientos,
        maximo=maximo,
        eje_y=eje_y,
        filas_deudas=filas_deudas,
        maximo_deudas=maximo_deudas,
        eje_y_deudas=eje_y_deudas,
        totales=totales,
    )


@app.route("/resumen")
def resumen():
    todos_movimientos = leer_movimientos()
    ciclos = opciones_ciclos(todos_movimientos)
    ciclo_seleccionado = request.args.get("ciclo") or (
        ciclos[0]["valor"] if ciclos else clave_ciclo(fecha_hoy_chile(), todos_movimientos)
    )
    if ciclos and ciclo_seleccionado not in {item["valor"] for item in ciclos}:
        ciclo_seleccionado = ciclos[0]["valor"]
    ciclo_etiqueta = next(
        (item["etiqueta"] for item in ciclos if item["valor"] == ciclo_seleccionado),
        etiqueta_ciclo(ciclo_seleccionado),
    )
    ciclo_tiene_informacion = next(
        (
            item["tiene_informacion"]
            for item in ciclos
            if item["valor"] == ciclo_seleccionado
        ),
        True,
    )
    movimientos, periodo_inicio, periodo_fin = movimientos_de_ciclo(
        todos_movimientos, ciclo_seleccionado
    )
    busqueda = request.args.get("q", "").strip()
    filtro_tipo = request.args.get("tipo", "Todos")
    semanas, _, _ = calcular_semanas_restantes(fecha_hoy_chile(), periodo_fin)
    ingresos = sum(item["monto"] for item in movimientos if item["tipo"] == "Ingreso")
    gastos = sum(item["monto"] for item in movimientos if item["tipo"] == "Gasto")
    ahorros = sum(item["monto"] for item in movimientos if item["tipo"] == "Ahorro")
    disponible = ingresos - gastos - ahorros
    cuota_semanal = disponible / semanas if semanas > 0 else 0
    totales = {}
    for item in movimientos:
        clave = (item["tipo"], item["categoria"])
        totales[clave] = totales.get(clave, 0) + item["monto"]
    por_categoria = [
        {"tipo": tipo, "categoria": categoria, "monto": monto}
        for (tipo, categoria), monto in totales.items()
    ]
    por_categoria.sort(key=lambda item: (item["tipo"], -item["monto"]))

    gastos_por_categoria_base = [
        item for item in por_categoria if item["tipo"] == "Gasto" and item["monto"] > 0
    ]
    gastos_por_categoria = gastos_por_categoria_base[:7]
    otros_gastos = gastos_por_categoria_base[7:]
    if otros_gastos:
        gastos_por_categoria.append(
            {
                "tipo": "Gasto",
                "categoria": "Otros",
                "monto": sum(item["monto"] for item in otros_gastos),
            }
        )
    gastos_por_categoria, torta_gastos, total_gastos_categoria = preparar_segmentos(
        gastos_por_categoria
    )

    deuda = sum(
        item["monto"]
        for item in movimientos
        if item.get("tipo") == "Gasto" and "deuda" in (item.get("categoria") or "").lower()
    )
    costos_sin_deuda = max(gastos - deuda, 0)
    disponible_ingresos = max(ingresos - gastos - ahorros, 0)
    distribucion_ingresos = [
        {"categoria": "Gastos", "monto": costos_sin_deuda, "color": "#1f6feb"},
        {"categoria": "Deuda", "monto": deuda, "color": "#c2410c"},
        {"categoria": "Ahorro", "monto": ahorros, "color": "#7c3aed"},
        {"categoria": "Disponible", "monto": disponible_ingresos, "color": "#15803d"},
    ]
    distribucion_ingresos = [item for item in distribucion_ingresos if item["monto"] > 0]
    distribucion_ingresos, torta_ingresos, total_distribucion_ingresos = preparar_segmentos(
        distribucion_ingresos
    )

    movimientos_filtrados = movimientos
    if filtro_tipo in TIPOS_VALIDOS:
        movimientos_filtrados = [
            item for item in movimientos_filtrados if item["tipo"] == filtro_tipo
        ]
    if busqueda:
        texto = busqueda.lower()
        movimientos_filtrados = [
            item
            for item in movimientos_filtrados
            if texto
            in " ".join(
                [
                    item.get("fecha", ""),
                    item.get("tipo", ""),
                    item.get("categoria", ""),
                    item.get("descripcion", ""),
                    str(item.get("monto", "")),
                ]
            ).lower()
        ]

    return render_template(
        "resumen.html",
        ingresos=ingresos,
        gastos=gastos,
        ahorros=ahorros,
        balance=ingresos - gastos,
        disponible=disponible,
        cuota_semanal=cuota_semanal,
        periodo_inicio=periodo_inicio,
        periodo_fin=periodo_fin,
        ciclos=ciclos,
        ciclo_seleccionado=ciclo_seleccionado,
        ciclo_etiqueta=ciclo_etiqueta,
        ciclo_tiene_informacion=ciclo_tiene_informacion,
        por_categoria=por_categoria,
        gastos_por_categoria=gastos_por_categoria,
        torta_gastos=torta_gastos,
        total_gastos_categoria=total_gastos_categoria,
        distribucion_ingresos=distribucion_ingresos,
        torta_ingresos=torta_ingresos,
        total_distribucion_ingresos=total_distribucion_ingresos,
        movimientos=movimientos,
        movimientos_filtrados=movimientos_filtrados,
        busqueda=busqueda,
        filtro_tipo=filtro_tipo,
    )


@app.route("/resumen/exportar.xlsx")
def exportar_movimientos_ciclo():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    todos_movimientos = leer_movimientos()
    ciclo = request.args.get("ciclo", "").strip()
    ciclos_validos = {item["valor"] for item in opciones_ciclos(todos_movimientos)}
    if ciclo not in ciclos_validos:
        ciclo = clave_ciclo(fecha_hoy_chile(), todos_movimientos)
    movimientos, inicio, fin = movimientos_de_ciclo(todos_movimientos, ciclo)

    libro = Workbook()
    hoja = libro.active
    hoja.title = "Movimientos"
    encabezados = ["Fecha", "Tipo", "Categoría", "Descripción", "Monto"]
    hoja.append(encabezados)
    for celda in hoja[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="1F6FEB")

    for item in sorted(movimientos, key=lambda movimiento: movimiento.get("fecha", "")):
        fecha = fecha_movimiento(item.get("fecha", ""))
        hoja.append(
            [
                fecha,
                item.get("tipo", ""),
                item.get("categoria", "") or CATEGORIA_SIN_ASIGNAR,
                item.get("descripcion", ""),
                float(item.get("monto") or 0),
            ]
        )
    for celda in hoja["A"][1:]:
        celda.number_format = "dd-mm-yyyy"
    for celda in hoja["E"][1:]:
        celda.number_format = '$' + '#,##0.00'
    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = hoja.dimensions
    for columna, ancho in {"A": 14, "B": 14, "C": 24, "D": 38, "E": 18}.items():
        hoja.column_dimensions[columna].width = ancho

    salida = BytesIO()
    libro.save(salida)
    salida.seek(0)
    nombre = f"movimientos_{inicio.isoformat()}_{fin.isoformat()}.xlsx"
    return send_file(
        salida,
        as_attachment=True,
        download_name=nombre,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    asegurar_csv()
    app.run(debug=True)
